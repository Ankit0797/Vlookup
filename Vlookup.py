import openpyxl
 
daily_data = openpyxl.load_workbook('C:/Users/ankit/Desktop/new.xlsx')
master_data = openpyxl.load_workbook('C:/Users/ankit/Desktop/hello.xlsx')
 
daily_sheet = daily_data['Sheet1']
master_sheet = master_data['Sheet1']
 
for i in daily_sheet.iter_rows():
    id = i[0].value
    print(id)
    row_number = i[0].row
    for j in master_sheet.iter_rows():
        print(j[0].value)
        if j[0].value == id:
            print(j[0].value)
            daily_sheet.cell(row=row_number, column=4).value = j[1].value
            daily_sheet.cell(row=row_number, column=5).value = j[2].value
            daily_sheet.cell(row=row_number, column=6).value = j[3].value
 
daily_data.save('C:/Users/ankit/Desktop/update_daily_sheet.xlsx')